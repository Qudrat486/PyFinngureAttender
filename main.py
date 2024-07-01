import os
import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QFileDialog, QTextEdit, QMessageBox
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor, QCursor, QIcon

class AttendanceLogApp(QWidget):
    noFoundIndex = list()
    path = str()
    def __init__(self):
        super().__init__()

        self.setWindowTitle("PyFingureAttender")
        self.setWindowIcon(QIcon(resource_path('.\\assets\\icon.ico')))
        self.setGeometry(100, 100, 400,430)
        
        
        # Set maximum size for the window
        self.setMaximumSize(700, 700)
        self.setStyleSheet("background-color: #F0F0F0;")
        self.init_ui()

    def init_ui(self):
        self.info_label = QLabel(self)
        self.info_label.setText("PyFingureAttender")
        self.info_label.setGeometry(30, 10, 350, 50)
        # Style for the label
        style = (
            "color: black; font-weight: bold; font-size: 20px; font-family: Arial, sans-serif;"
        )

        self.info_label.setStyleSheet(style)
        
        # stylesheets
        txtbox_styleSheet = "background-color: #F0F0F0; color: black; font-size: 12px; border-radius: 5px;border-width: 1px; border-style: solid; border-color: #005CA5;"
        button_StyleSheet = "QPushButton:hover { background-color: #005CA5;} QPushButton { background-color: #0078D4; color: white; border-radius: 5px;font-weight: bold; font-size: 12px;}"
        
        
        self.file_entry = QLineEdit(self)
        self.file_entry.setGeometry(30, 80, 250, 30)
        self.file_entry.setStyleSheet(txtbox_styleSheet)
        
        self.browse_button = QPushButton("Browse", self)
        self.browse_button.setGeometry(285, 80, 90, 30)
        # self.browse_button.setIcon(QIcon("assets/folder-open.png"))
        self.browse_button.clicked.connect(self.browse_file)

        self.execute_button = QPushButton("Run", self)
        self.execute_button.setGeometry(30, 120, 171, 30)
        self.execute_button.clicked.connect(self.execute_script)

        self.reset_button = QPushButton("Reset", self)
        self.reset_button.setGeometry(206, 120, 171, 30)
        self.reset_button.clicked.connect(self.reset_fields)

        self.save_button = QPushButton("Save", self)
        self.save_button.setGeometry(285, 370, 90, 30)
        self.save_button.clicked.connect(self.save_output)
        
        self.message_box = QTextEdit(self)
        self.message_box.setGeometry(30, 160, 348, 200)
        self.message_box.setReadOnly(True)
        self.message_box.setStyleSheet(txtbox_styleSheet)

        # Set button colors using button_stylesheet
        self.browse_button.setStyleSheet(button_StyleSheet)
        self.browse_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.execute_button.setStyleSheet(button_StyleSheet)
        self.execute_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.reset_button.setStyleSheet(button_StyleSheet)
        self.reset_button.setCursor(QCursor(Qt.PointingHandCursor))
        self.save_button.setStyleSheet(button_StyleSheet)
        self.save_button.setCursor(QCursor(Qt.PointingHandCursor))
        
        # Copyright label
        copyright_label = QLabel(self)
        copyright_label.setText("© 2024 Qudrat Ullah(017-BSCS-21). All rights reserved.")
        copyright_label.setGeometry(30, 405, 350, 20)

        # Style for the copyright label
        copyright_style = "color: gray; font-size: 10px;"

        copyright_label.setStyleSheet(copyright_style)
        
    def browse_file(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Open DAT File", "", "DAT Files (*.dat)")
            if file_path:
                self.file_entry.setText(file_path)
        except Exception as e:
            self.message_box.append(f"{e}")
            
    def reset_fields(self):
        answer = QMessageBox.question(
            self,
            "Confirmation",
            "Do you want to reset?",
            QMessageBox.StandardButton.Yes |
            QMessageBox.StandardButton.No
        )

        if answer == QMessageBox.StandardButton.Yes:
            self.file_entry.clear()
            self.message_box.clear()

    def save_output(self, notfoundedindex):
     with open(self.path, "w") as fhand:
         for index in self.noFoundIndex:
             fhand.write(index + "\n")
         QMessageBox.information(
             self,
             'Information',
             'Data saved successfully!'
         )

        
    
    def execute_script(self):
        try:
            file_path = self.file_entry.text()
            output_file_path = file_path.split('.')[0]
            self.path = output_file_path + ".txt"
            current_index, month_input = self.extract_date(file_path)

            # end_index = 200 if current_index == 101 else 350
            end_index = 500
            QMessageBox.information(
             self,
             'Processing...',
             'Keep patience for a while!'
         )
            while current_index <= end_index:
                data_list = self.process_attendance_log(current_index, file_path)
                result_dict = self.extract_data(current_index, data_list)

                if str(current_index) in result_dict:
                    date_ranges = self.find_date_ranges(result_dict)
                    result_dict = self.fill_missing_dates(result_dict, date_ranges)
                    sorted_data = {k: dict(sorted(v.items(), key=lambda x: x[0])) for k, v in result_dict[str(current_index)].items()}
                    sorted_data = dict(sorted(sorted_data.items(), key=lambda x: x[0]))
                    s = current_index
                    data_dict = {}
                    data_dict[str(s)] = sorted_data
                    self.create_xlsx(current_index, data_dict, month_input,output_file_path)
                else:
                    self.message_box.append(f"No data found for index {current_index}")
                    self.noFoundIndex.append(f"No data found for index {current_index}")
                current_index += 1

            self.message_box.append("XLSX file created successfully.")
        except Exception as e:
            # self.message_box.append(f"{e}")
            QMessageBox.warning(
             self,
             'Error',
             f'{e}'
         )
 
    def process_attendance_log(self, current_index, file_path):
        filtered_data = []

        with open(file_path) as attendance_log:
            attendance_data = [line.replace("\t", " ").replace("\n", " ").strip() for line in attendance_log]

        filtered = [line for line in attendance_data if line.startswith(str(current_index))]

        for i in range(len(filtered)):
            current_timestamp = filtered[i][12:17]  # Extract timestamp (assuming timestamps are at positions 12:17)

            if i < len(filtered) - 1 and current_timestamp == filtered[i + 1][12:17]:
                # Check if the next element has the same timestamp
                if not filtered_data or filtered_data[-1][12:17] != current_timestamp:
                    filtered_data.append(filtered[i])
            elif not filtered_data or filtered_data[-1][12:17] != current_timestamp:
                # If filtered_data is empty or the timestamps don't match, add the current element to filtered_data
                filtered_data.append(filtered[i])

        return filtered_data

    def extract_date(self, file_path):
        index = None
        month = None
        with open(file_path) as attendance_log:
            for line in attendance_log:
                line = line.replace("\t", " ").replace("\n", " ").strip()
                index, month = tuple(line.split()[0:2])
                break
        index, month = 101 if index[0] == "1" else 201, int(month.split("-")[1])
        index = 1
        return tuple([index, month])

    def extract_data(self, current_index, data_list):
        result_dict = {}
        user_id = str(current_index)

        for item in data_list:
            parts = item.split()

            # Check if parts has at least three elements before trying to access its elements
            if len(parts) >= 3:
                user_id = parts[0]
                date = parts[1]
                time = int(parts[2].split(":")[0])

                if user_id not in result_dict:
                    result_dict[user_id] = {}

                if time < 12:
                    result_dict[user_id][date] = {"checkIn": parts[2], "checkOut": ""}
                else:
                    if date not in result_dict[user_id]:
                        result_dict[user_id][date] = {"checkIn": "", "checkOut": parts[2]}
                    else:
                        result_dict[user_id][date]["checkOut"] = parts[2]
            else:
                if user_id not in result_dict:
                    result_dict[user_id] = {}
                result_dict[user_id][date] = {"checkIn": "", "checkOut": ""}
        return result_dict

    def fill_missing_dates(self, result_dict, date_ranges):
        for user_id, dates in date_ranges.items():
            if len(dates) >= 1:
                earliest_date, latest_date = dates[0], dates[-1]

                # Extract the year and month from the earliest_date
                year, month, _ = map(int, earliest_date.split('-'))

                # Get the first day of the month
                start_date = datetime(year, month, 1)

                # Get the last day of the month
                end_date = datetime(year, month % 12 + 1, 1) - timedelta(days=1)

                # Generate all dates for the given month
                all_dates = [str(start_date + timedelta(days=i)).split()[0] for i in range((end_date - start_date).days + 1)]

                for date in all_dates:
                    if date not in result_dict[user_id]:
                        result_dict[user_id][date] = {"checkIn": "", "checkOut": ""}
        return result_dict

    def find_date_ranges(self, result_dict):
        return {user_id: sorted(result_dict[user_id].keys()) for user_id in result_dict}

    def create_xlsx(self, index, result_dict, month_input,output_file_path):
        start_date = datetime(datetime.now().year, month_input, 1)
        end_date = datetime(datetime.now().year, month_input % 12 + 1, 1) - timedelta(
            days=1) if month_input < 12 else datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)
        output_file_path = f"{output_file_path}.xlsx"

        if os.path.exists(output_file_path):
            workbook = load_workbook(output_file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active

            # Writing the initial header row with merged cells for dates
            header_row = ['Index']
            for date in pd.date_range(start=start_date, end=end_date, freq='D'):
                header_row.append(date.strftime('%Y-%m-%d'))
                header_row.append('')  # Empty cell for merging
            sheet.append(header_row)

        # Get existing dates from the header row
        existing_dates = [cell.value for cell in sheet[1]]
        n = len(existing_dates)
        # print(existing_dates)
        # exit()
        # Writing the data
        for user_id, user_data in result_dict.items():
            data_row = [user_id]

            for date, times in user_data.items():
                if date in existing_dates:
                    # Find the column index for the existing date
                    col_index = existing_dates.index(date) + 1
                    data_row.extend([times['checkIn'], times['checkOut']])
                else:
                    # Add a new column for the new date in the header row
                    sheet.cell(row=1, column=len(existing_dates) + 1, value=date.strftime('%Y-%m-%d'))
                    sheet.cell(row=1, column=len(existing_dates) + 2, value='')  # Empty cell for merging
                    data_row.extend(['', ''])  # Add empty values for previous dates
                    data_row.extend(['', ''])  # Add empty values for the new date

            # Find the row index for the user_id
            row_index = len(sheet['A']) + 1
            # Write the data to the correct columns
            for col_index, value in enumerate(data_row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)


        workbook.save(output_file_path)

   
    
if __name__ == "__main__":
    def resource_path(relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS2
            base_path = sys._MEIPASS2
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    app = QApplication([])
    attendance_app = AttendanceLogApp()
    attendance_app.show()
    app.exec_()
